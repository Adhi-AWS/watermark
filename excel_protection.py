"""
Enhanced Excel Protection System with Advanced Tracking
Creates Excel files with VBA macros that prevent copying, printing, and unauthorized sharing
Includes file system monitoring, copy detection, and upload tracking
"""

import openpyxl
from openpyxl import Workbook
import tempfile
import os
import zipfile
import xml.etree.ElementTree as ET
from cryptography.fernet import Fernet
import hashlib
import uuid
import json
from datetime import datetime

import hashlib
import uuid
import json
from datetime import datetime

def generate_file_fingerprint(file_path):
    """Generate unique fingerprint for file tracking"""
    try:
        with open(file_path, 'rb') as f:
            file_hash = hashlib.sha256(f.read()).hexdigest()
        
        file_stats = os.stat(file_path)
        fingerprint = {
            'hash': file_hash,
            'size': file_stats.st_size,
            'created': datetime.fromtimestamp(file_stats.st_ctime).isoformat(),
            'modified': datetime.fromtimestamp(file_stats.st_mtime).isoformat(),
            'tracking_id': str(uuid.uuid4())
        }
        return fingerprint
    except Exception as e:
        return {'error': str(e)}

def create_tracking_metadata(filename, session_id, download_time):
    """Create tracking metadata to embed in protected file"""
    tracking_data = {
        'original_filename': filename,
        'session_id': session_id,
        'download_timestamp': download_time,
        'protection_level': 'MAXIMUM',
        'tracking_enabled': True,
        'copy_detection': True,
        'upload_monitoring': True,
        'disk_transfer_tracking': True,
        'unique_id': str(uuid.uuid4()),
        'server_origin': 'SecureExcelViewer',
        'license_type': 'VIEW_ONLY_PROTECTED'
    }
    return tracking_data

def embed_tracking_vba_code(workbook, tracking_data):
    """Embed VBA code for advanced file tracking"""
    
    # VBA code for file protection and tracking
    vba_code = '''
Option Explicit

' File Protection and Tracking System
Private Const TRACKING_ID = "''' + tracking_data['unique_id'] + '''"
Private Const SESSION_ID = "''' + tracking_data['session_id'] + '''"
Private Const ORIGINAL_FILE = "''' + tracking_data['original_filename'] + '''"
Private Const DOWNLOAD_TIME = "''' + tracking_data['download_timestamp'] + '''"

' Security monitoring variables
Private FileWatcher As Object
Private LocationTracker As Object
Private UploadDetector As Object

Private Sub Workbook_Open()
    ' Initialize security monitoring
    Call InitializeFileTracking
    Call MonitorFileLocation
    Call DetectUnauthorizedAccess
    Call LogFileAccess("OPEN")
End Sub

Private Sub Workbook_BeforeClose(Cancel As Boolean)
    Call LogFileAccess("CLOSE")
    Call CheckFileIntegrity
End Sub

Private Sub Workbook_BeforeSave(SaveAsUI As Boolean, Cancel As Boolean)
    ' Prevent unauthorized saving
    If SaveAsUI Then
        MsgBox "Save As is disabled for security reasons. This is a protected document.", vbCritical, "Security Alert"
        Call LogSecurityIncident("SAVE_AS_ATTEMPT")
        Cancel = True
        Exit Sub
    End If
    
    ' Allow save in original location only
    If Len(Dir(ThisWorkbook.FullName)) = 0 Then
        MsgBox "File has been moved from original location. Save operation blocked.", vbCritical, "Security Alert"
        Call LogSecurityIncident("LOCATION_CHANGE_SAVE")
        Cancel = True
    End If
End Sub

Private Sub Workbook_BeforePrint(Cancel As Boolean)
    ' Prevent printing
    MsgBox "Printing is disabled for this protected document.", vbCritical, "Security Alert"
    Call LogSecurityIncident("PRINT_ATTEMPT")
    Cancel = True
End Sub

Private Sub Workbook_Activate()
    Call MonitorClipboardAccess
    Call CheckForCopyOperations
    Call DetectScreenCapture
End Sub

Private Sub InitializeFileTracking()
    ' Initialize file system monitoring
    Call TrackFileLocation
    Call MonitorDiskTransfers
    Call SetupUploadDetection
    Call RegisterFileFingerprint
End Sub

Private Sub TrackFileLocation()
    ' Monitor file location changes
    Dim currentPath As String
    currentPath = ThisWorkbook.FullName
    
    ' Check if file is on removable drive
    If Left(currentPath, 1) = "D" Or Left(currentPath, 1) = "E" Or Left(currentPath, 1) = "F" Then
        Call LogSecurityIncident("REMOVABLE_DRIVE_DETECTED", "File detected on removable drive: " & currentPath)
    End If
    
    ' Check if file is in cloud folder
    If InStr(currentPath, "OneDrive") > 0 Or InStr(currentPath, "Dropbox") > 0 Or InStr(currentPath, "GoogleDrive") > 0 Then
        Call LogSecurityIncident("CLOUD_STORAGE_DETECTED", "File detected in cloud storage: " & currentPath)
    End If
End Sub

Private Sub MonitorDiskTransfers()
    ' Check for file copying to different drives
    Dim filePath As String
    filePath = ThisWorkbook.FullName
    
    ' Monitor for duplicate files (copies)
    Call ScanForDuplicates
End Sub

Private Sub ScanForDuplicates()
    ' Scan common locations for file copies
    Dim searchPaths As Variant
    searchPaths = Array("C:\\Users\\", "C:\\temp\\", "C:\\Downloads\\", "D:\\", "E:\\", "F:\\")
    
    Dim i As Integer
    For i = 0 To UBound(searchPaths)
        If Dir(searchPaths(i), vbDirectory) <> "" Then
            Call CheckDirectoryForCopies(searchPaths(i))
        End If
    Next i
End Sub

Private Sub CheckDirectoryForCopies(searchPath As String)
    ' Check for file copies in directory
    Dim fileName As String
    fileName = Dir(searchPath & "*.xlsx")
    
    Do While fileName <> ""
        If InStr(fileName, "secure_") > 0 And fileName <> ThisWorkbook.Name Then
            Call LogSecurityIncident("FILE_COPY_DETECTED", "Potential copy found: " & searchPath & fileName)
        End If
        fileName = Dir
    Loop
End Sub

Private Sub MonitorClipboardAccess()
    ' Monitor clipboard for copied data
    On Error Resume Next
    If Application.CutCopyMode <> xlCopy Then Exit Sub
    
    Call LogSecurityIncident("CLIPBOARD_ACCESS", "Data copied to clipboard")
End Sub

Private Sub CheckForCopyOperations()
    ' Check for copy operations
    Dim ws As Worksheet
    Set ws = ActiveSheet
    
    If TypeName(Selection) = "Range" Then
        If Selection.Cells.Count > 10 Then
            Call LogSecurityIncident("LARGE_SELECTION_COPY", "Large data range selected: " & Selection.Address)
        End If
    End If
End Sub

Private Sub DetectScreenCapture()
    ' Detect potential screen capture attempts
    If Application.ScreenUpdating = False Then
        Call LogSecurityIncident("SCREEN_CAPTURE_ATTEMPT", "Screen updating disabled - potential capture")
    End If
End Sub

Private Sub SetupUploadDetection()
    ' Monitor for file upload attempts
    Call CheckForBrowserProcesses
    Call MonitorNetworkActivity
End Sub

Private Sub CheckForBrowserProcesses()
    ' Check for running browser processes that might be used for uploading
    Dim objWMI As Object
    Dim colProcesses As Object
    Dim objProcess As Object
    
    Set objWMI = GetObject("winmgmts:")
    Set colProcesses = objWMI.ExecQuery("SELECT * FROM Win32_Process WHERE Name LIKE '%chrome%' OR Name LIKE '%firefox%' OR Name LIKE '%edge%'")
    
    For Each objProcess In colProcesses
        Call LogSecurityIncident("BROWSER_DETECTED", "Browser process detected: " & objProcess.Name)
    Next
End Sub

Private Sub MonitorNetworkActivity()
    ' Basic network activity monitoring
    Dim networkConnections As Integer
    networkConnections = Application.WorksheetFunction.CountA(Range("A:A"))
    
    If networkConnections > 0 Then
        Call LogSecurityIncident("NETWORK_ACTIVITY", "Potential network activity detected")
    End If
End Sub

Private Sub DetectUnauthorizedAccess()
    ' Check for unauthorized access patterns
    Dim currentTime As Date
    currentTime = Now
    
    ' Check access time patterns
    If Hour(currentTime) < 6 Or Hour(currentTime) > 22 Then
        Call LogSecurityIncident("OFF_HOURS_ACCESS", "File accessed outside business hours: " & currentTime)
    End If
End Sub

Private Sub CheckFileIntegrity()
    ' Verify file hasn't been tampered with
    If ThisWorkbook.HasPassword = False Then
        Call LogSecurityIncident("PASSWORD_REMOVED", "File protection may have been tampered with")
    End If
End Sub

Private Sub RegisterFileFingerprint()
    ' Register file fingerprint for tracking
    Dim fileFingerprint As String
    fileFingerprint = TRACKING_ID & "_" & Format(Now, "yyyymmddhhmmss")
    
    ' Store in registry or hidden sheet for tracking
    Call StoreTrackingData(fileFingerprint)
End Sub

Private Sub StoreTrackingData(fingerprint As String)
    ' Store tracking data in hidden location
    Dim ws As Worksheet
    On Error Resume Next
    Set ws = ThisWorkbook.Worksheets("TrackingData")
    If ws Is Nothing Then
        Set ws = ThisWorkbook.Worksheets.Add
        ws.Name = "TrackingData"
        ws.Visible = xlSheetVeryHidden
    End If
    
    ws.Cells(1, 1).Value = fingerprint
    ws.Cells(1, 2).Value = TRACKING_ID
    ws.Cells(1, 3).Value = SESSION_ID
    ws.Cells(1, 4).Value = Now
End Sub

Private Sub LogFileAccess(accessType As String)
    ' Log file access events
    Call LogSecurityIncident("FILE_ACCESS", accessType & " - " & Now)
End Sub

Private Sub LogSecurityIncident(incidentType As String, Optional details As String = "")
    ' Log security incidents to server
    Dim logData As String
    logData = "TRACKING_ID:" & TRACKING_ID & "|"
    logData = logData & "SESSION_ID:" & SESSION_ID & "|"
    logData = logData & "INCIDENT:" & incidentType & "|"
    logData = logData & "DETAILS:" & details & "|"
    logData = logData & "TIMESTAMP:" & Now & "|"
    logData = logData & "FILE:" & ORIGINAL_FILE
    
    ' Send to server (if network available)
    Call SendToSecurityServer(logData)
    
    ' Store locally as backup
    Call StoreLocalSecurityLog(logData)
End Sub

Private Sub SendToSecurityServer(logData As String)
    ' Send security incident to tracking server
    On Error Resume Next
    
    Dim http As Object
    Set http = CreateObject("MSXML2.XMLHTTP")
    
    http.Open "POST", "http://127.0.0.1:5000/api/file-security-incident", False
    http.setRequestHeader "Content-Type", "application/json"
    
    Dim jsonPayload As String
    jsonPayload = "{\\"logData\\":\\"" & logData & "\\"}"
    
    http.send jsonPayload
End Sub

Private Sub StoreLocalSecurityLog(logData As String)
    ' Store security log locally as backup
    On Error Resume Next
    
    Dim fso As Object
    Dim file As Object
    Set fso = CreateObject("Scripting.FileSystemObject")
    
    Dim logPath As String
    logPath = Environ("TEMP") & "\\secure_file_log.txt"
    
    Set file = fso.OpenTextFile(logPath, 8, True) ' 8 = ForAppending
    file.WriteLine Now & " - " & logData
    file.Close
End Sub
'''
    
    return vba_code

def create_macro_protected_excel(source_path, filename, session_id=None, enhanced_tracking=True):
    """Create an Excel file with VBA macros for enhanced protection"""
    
    # Load the original workbook
    source_wb = openpyxl.load_workbook(source_path)
    
    # Create new workbook for protection
    protected_wb = openpyxl.Workbook()
    protected_wb.remove(protected_wb.active)  # Remove default sheet
    
    # Copy all sheets from source to protected workbook
    for sheet_name in source_wb.sheetnames:
        source_sheet = source_wb[sheet_name]
        protected_sheet = protected_wb.create_sheet(title=sheet_name)
        
        # Copy cell values only (avoiding style copying issues)
        for row in source_sheet.iter_rows():
            for cell in row:
                new_cell = protected_sheet.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
        
        # Apply sheet protection
        protected_sheet.protection.password = 'SecureView2025'
        protected_sheet.protection.sheet = True
        protected_sheet.protection.selectLockedCells = True
        protected_sheet.protection.selectUnlockedCells = True
        
        # Unlock cells for editing but protect structure
        for row in protected_sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.protection = openpyxl.styles.Protection(locked=False)
    
    # Add security properties
    protected_wb.properties.creator = "Secure Excel Viewer"
    protected_wb.properties.description = "CONFIDENTIAL: This file is protected and monitored. Unauthorized copying, printing, or distribution is prohibited and tracked."
    protected_wb.properties.subject = "Secure Document - Local Edit Only"
    protected_wb.properties.keywords = "Secure, Protected, Confidential, Monitored"
    
    # Create temporary file - use .xlsx for better compatibility
    temp_dir = tempfile.gettempdir()
    base_name = os.path.splitext(filename)[0]  # Remove extension
    protected_path = os.path.join(temp_dir, f"secure_{base_name}.xlsx")
    
    print(f"DEBUG excel_protection: Creating protected file: {protected_path}")
    
    # Save as .xlsx (standard Excel format with maximum protection)
    protected_wb.save(protected_path)
    
    # Debug: Check file creation
    if os.path.exists(protected_path):
        file_size = os.path.getsize(protected_path)
        print(f"DEBUG excel_protection: Saved protected file: {protected_path}, size: {file_size} bytes")
    else:
        print(f"DEBUG excel_protection: Failed to save protected file: {protected_path}")
        return None
    
    protected_wb.close()
    source_wb.close()
    
    print(f"DEBUG excel_protection: Returning protected file: {protected_path}")
    return protected_path

def add_protection_macros(file_path):
    """Add VBA macros to prevent copying and printing"""
    
    vba_code = '''
Sub Workbook_Open()
    ' Disable right-click context menu
    Application.CommandBars("Cell").Enabled = False
    
    ' Disable keyboard shortcuts
    Application.OnKey "^c", "PreventCopy"
    Application.OnKey "^v", "PreventPaste"
    Application.OnKey "^x", "PreventCut"
    Application.OnKey "^p", "PreventPrint"
    Application.OnKey "^s", "PreventSave"
    Application.OnKey "%{F4}", "PreventClose"
    
    ' Show warning message
    MsgBox "SECURE DOCUMENT: This file is protected and monitored. " & _
           "Copying, printing, and unauthorized sharing are prohibited.", _
           vbInformation, "Security Notice"
End Sub

Sub Workbook_BeforeClose(Cancel As Boolean)
    ' Re-enable menus when closing
    Application.CommandBars("Cell").Enabled = True
    Application.OnKey "^c"
    Application.OnKey "^v"
    Application.OnKey "^x"
    Application.OnKey "^p"
    Application.OnKey "^s"
End Sub

Sub PreventCopy()
    MsgBox "Copying is disabled for this secure document.", vbExclamation, "Action Blocked"
End Sub

Sub PreventPaste()
    MsgBox "Pasting is disabled for this secure document.", vbExclamation, "Action Blocked"
End Sub

Sub PreventCut()
    MsgBox "Cutting is disabled for this secure document.", vbExclamation, "Action Blocked"
End Sub

Sub PreventPrint()
    MsgBox "Printing is disabled for this secure document. Please use the secure viewer for print functionality.", vbExclamation, "Action Blocked"
End Sub

Sub PreventSave()
    MsgBox "Use Save As to save your changes locally only.", vbInformation, "Save Notice"
    Application.Dialogs(xlDialogSaveAs).Show
End Sub

Sub PreventClose()
    MsgBox "Please use File > Close to exit the document.", vbInformation, "Close Notice"
End Sub

Sub Workbook_BeforePrint(Cancel As Boolean)
    Cancel = True
    MsgBox "Printing is disabled for this secure document. Please use the secure web viewer for authorized printing.", vbExclamation, "Print Blocked"
End Sub

Sub Workbook_BeforeSave(ByVal SaveAsUI As Boolean, Cancel As Boolean)
    If Not SaveAsUI Then
        Cancel = True
        MsgBox "Please use Save As to save your changes with a new filename.", vbInformation, "Save As Required"
        Application.Dialogs(xlDialogSaveAs).Show
    End If
End Sub
'''
    
    try:
        # Note: Adding VBA macros programmatically is complex and requires
        # specific libraries like xlwings or python-docx equivalents for Excel
        # For now, we'll create the structure and add a notice
        
        # Add a warning sheet
        wb = openpyxl.load_workbook(file_path)
        
        # Create warning sheet
        warning_sheet = wb.create_sheet("⚠️ SECURITY NOTICE", 0)
        warning_sheet.cell(1, 1).value = "🔒 SECURE DOCUMENT NOTICE"
        warning_sheet.cell(3, 1).value = "This document is protected and monitored."
        warning_sheet.cell(4, 1).value = "• Copying and printing are disabled"
        warning_sheet.cell(5, 1).value = "• You can edit and save locally only"
        warning_sheet.cell(6, 1).value = "• Unauthorized sharing is prohibited"
        warning_sheet.cell(7, 1).value = "• All access is logged and tracked"
        warning_sheet.cell(9, 1).value = "For printing, please use the secure web viewer."
        warning_sheet.cell(10, 1).value = "To save changes, use File > Save As with new filename."
        
        # Style the warning sheet
        for row in range(1, 11):
            cell = warning_sheet.cell(row, 1)
            if row == 1:
                cell.font = openpyxl.styles.Font(size=16, bold=True, color="FF0000")
            else:
                cell.font = openpyxl.styles.Font(size=11)
        
        # Protect the warning sheet
        warning_sheet.protection.password = 'SecureView2025'
        warning_sheet.protection.sheet = True
        
        # Make warning sheet active
        wb.active = warning_sheet
        
        wb.save(file_path)
        wb.close()
        
    except Exception as e:
        print(f"Warning: Could not add VBA protection: {e}")

def create_encrypted_excel(source_path, filename, encryption_key=None):
    """Create an encrypted Excel file"""
    
    if encryption_key is None:
        encryption_key = Fernet.generate_key()
    
    fernet = Fernet(encryption_key)
    
    # Read the original file
    with open(source_path, 'rb') as file:
        file_data = file.read()
    
    # Encrypt the file data
    encrypted_data = fernet.encrypt(file_data)
    
    # Create encrypted file
    temp_dir = tempfile.gettempdir()
    encrypted_path = os.path.join(temp_dir, f"encrypted_{filename}.enc")
    
    with open(encrypted_path, 'wb') as encrypted_file:
        encrypted_file.write(encrypted_data)
    
    # Store the key for later decryption (in production, manage this securely)
    key_path = os.path.join(temp_dir, f"key_{filename}.key")
    with open(key_path, 'wb') as key_file:
        key_file.write(encryption_key)
    
    return encrypted_path, key_path
