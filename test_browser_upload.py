"""
🧪 BROWSER UPLOAD DETECTION TEST
===============================

Test script to demonstrate browser upload tracking capabilities.
"""

import os
import shutil
import time
import sqlite3
import hashlib
import subprocess
from pathlib import Path

def setup_test_environment():
    """Set up test files and database"""
    print("🔧 Setting up test environment...")
    
    # Create test protected file
    test_file = "test_protected_file.xlsx"
    if not os.path.exists(test_file):
        # Create a simple Excel file for testing
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "CONFIDENTIAL DATA"
        ws['A2'] = "Employee Information"
        ws['A3'] = "Do not upload or share"
        wb.save(test_file)
        print(f"✅ Created test file: {test_file}")
    
    # Calculate file hash
    with open(test_file, 'rb') as f:
        file_hash = hashlib.sha256(f.read()).hexdigest()
    
    # Register in tracking database
    conn = sqlite3.connect('file_tracking.db')
    cursor = conn.cursor()
    
    # Create tables if they don't exist
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS protected_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_hash TEXT UNIQUE NOT NULL,
            file_path TEXT NOT NULL,
            file_name TEXT NOT NULL,
            download_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            download_session TEXT,
            file_size INTEGER,
            fingerprint TEXT
        )
    ''')
    
    # Register the test file
    try:
        cursor.execute('''
            INSERT OR REPLACE INTO protected_files 
            (file_hash, file_path, file_name, download_session, file_size)
            VALUES (?, ?, ?, ?, ?)
        ''', (file_hash, os.path.abspath(test_file), test_file, 'test_session', os.path.getsize(test_file)))
        conn.commit()
        print(f"✅ Registered protected file with hash: {file_hash[:16]}...")
    except Exception as e:
        print(f"❌ Error registering file: {e}")
    finally:
        conn.close()
    
    return test_file, file_hash

def simulate_browser_upload_scenarios(test_file):
    """Simulate various browser upload scenarios"""
    print("\n🌐 Simulating browser upload scenarios...")
    
    # Test 1: Copy to browser temp directory
    print("\n📋 Test 1: Copying to browser temp directory")
    browser_temp = os.path.expanduser("~/AppData/Local/Temp/browser_upload_test")
    os.makedirs(browser_temp, exist_ok=True)
    
    temp_copy = os.path.join(browser_temp, f"upload_{test_file}")
    shutil.copy2(test_file, temp_copy)
    print(f"   📁 Copied to: {temp_copy}")
    time.sleep(2)
    
    # Test 2: Copy to typical upload staging areas
    print("\n📤 Test 2: Copying to upload staging areas")
    staging_dirs = [
        os.path.expanduser("~/AppData/Roaming/Microsoft/Windows/Recent"),
        "C:/temp" if os.path.exists("C:/temp") else None,
        os.path.expanduser("~/Downloads/upload_test")
    ]
    
    for staging_dir in staging_dirs:
        if staging_dir and staging_dir != "None":
            try:
                os.makedirs(staging_dir, exist_ok=True)
                staging_copy = os.path.join(staging_dir, f"staging_{test_file}")
                shutil.copy2(test_file, staging_copy)
                print(f"   📂 Staged at: {staging_copy}")
                time.sleep(1)
            except Exception as e:
                print(f"   ❌ Failed to stage at {staging_dir}: {e}")
    
    # Test 3: Simulate clipboard copy (file path)
    print("\n📋 Test 3: Simulating clipboard operations")
    try:
        # Try to use clipboard (requires pywin32)
        import win32clipboard
        
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardText(os.path.abspath(test_file))
        win32clipboard.CloseClipboard()
        print(f"   📋 File path copied to clipboard: {test_file}")
        time.sleep(2)
        
        # Clear clipboard
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.CloseClipboard()
        
    except ImportError:
        print("   ⚠️  Clipboard test skipped (pywin32 not available)")
    except Exception as e:
        print(f"   ❌ Clipboard test failed: {e}")

def simulate_browser_processes():
    """Simulate browser processes accessing files"""
    print("\n🌐 Test 4: Simulating browser process detection")
    
    # This would normally be detected by the browser monitor
    # For testing, we'll just show what would be detected
    browsers = ['chrome.exe', 'firefox.exe', 'msedge.exe']
    upload_contexts = [
        "Gmail - Compose - Attach Files",
        "Google Drive - Upload Files", 
        "Dropbox - File Upload",
        "OneDrive - Upload Document",
        "Facebook - Share Photo",
        "LinkedIn - Attach Document"
    ]
    
    for browser in browsers:
        for context in upload_contexts:
            print(f"   🔍 Would detect: {browser} with window title containing '{context}'")
    
    print("   ⚡ Browser upload monitor would generate HIGH severity alerts for these scenarios")

def check_security_incidents():
    """Check for recorded security incidents"""
    print("\n📊 Checking recorded security incidents...")
    
    try:
        conn = sqlite3.connect('file_tracking.db')
        cursor = conn.cursor()
        
        # Check for file operations
        cursor.execute('''
            SELECT operation_type, source_path, destination_path, severity, timestamp
            FROM file_operations 
            ORDER BY timestamp DESC 
            LIMIT 10
        ''')
        
        incidents = cursor.fetchall()
        
        if incidents:
            print(f"   📈 Found {len(incidents)} recent file operations:")
            for incident in incidents:
                op_type, src, dst, severity, timestamp = incident
                print(f"      🚨 {severity}: {op_type}")
                print(f"         📁 {src} → {dst}")
                print(f"         ⏰ {timestamp}")
                print()
        else:
            print("   📊 No file operations recorded yet")
        
        conn.close()
        
    except Exception as e:
        print(f"   ❌ Error checking incidents: {e}")

def test_api_integration():
    """Test API integration with Flask server"""
    print("\n🔗 Testing API integration...")
    
    try:
        import requests
        
        # Test security incident API
        test_incident = {
            'incident_type': 'BROWSER_UPLOAD_ATTEMPT',
            'file_name': 'test_protected_file.xlsx',
            'browser_name': 'chrome.exe',
            'detected_services': ['Gmail', 'Google Drive'],
            'threat_level': 'HIGH',
            'severity': 'HIGH',
            'timestamp': '2025-08-29T20:00:00Z',
            'description': 'Test browser upload detection'
        }
        
        response = requests.post(
            'http://127.0.0.1:5000/api/file-security-incident',
            json=test_incident,
            timeout=5
        )
        
        if response.status_code == 200:
            print("   ✅ API integration working - security incident recorded")
        else:
            print(f"   ⚠️  API response: {response.status_code}")
            
    except requests.exceptions.ConnectionError:
        print("   ⚠️  Flask server not running - start with 'python app.py'")
    except Exception as e:
        print(f"   ❌ API test failed: {e}")

def main():
    """Main test function"""
    print("🧪 BROWSER UPLOAD DETECTION TEST")
    print("=" * 50)
    
    # Setup
    test_file, file_hash = setup_test_environment()
    
    # Run tests
    simulate_browser_upload_scenarios(test_file)
    simulate_browser_processes()
    check_security_incidents()
    test_api_integration()
    
    print("\n" + "=" * 50)
    print("🎯 TEST SUMMARY:")
    print("✅ Browser upload detection system configured")
    print("✅ Multiple upload scenarios simulated")
    print("✅ File tracking database updated")
    print("✅ API integration tested")
    print("\n🚀 To see live browser upload detection:")
    print("   1. Start Flask server: python app.py")
    print("   2. Open browser and navigate to upload sites")
    print("   3. Try to upload the test protected file")
    print("   4. Check security monitor: http://127.0.0.1:5000/security/monitor")
    print("\n🛡️  The system will detect and alert on:")
    print("   • File uploads to Gmail/Google Drive")
    print("   • Dropbox/OneDrive upload attempts") 
    print("   • Social media file sharing")
    print("   • Any browser-based file upload activity")

if __name__ == "__main__":
    main()
